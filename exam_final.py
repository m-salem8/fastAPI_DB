import openpyxl
from fastapi import FastAPI, HTTPException, Depends
from fastapi.security import HTTPBasic, HTTPBasicCredentials
from typing import Optional
from pydantic import BaseModel
import random



# Users database of name and coresponding password
users_db = {"alice": "wonderland",
            "bob": "builder",
            "clementine": "mandarine"}



# reading and loading the excel sheet
db = openpyxl.load_workbook('questions_en.xlsx')
sheet = db.active

# instantiating and API instance
api = FastAPI(title= "API to give exam to users", 
              description="API POWERED BY FASTAPI")

# Basic authentication
security = HTTPBasic()

#fuctin to verify the user
def verify_user(credentials: HTTPBasicCredentials = Depends(security)):
    username = credentials.username
    password = credentials.password
    
    if username in users_db and users_db[username] == password:
        return True
    else:
        raise HTTPException(status_code=401, detail="Unauthorized")
    
#function to verify the admin
def verify_admin(credentials: HTTPBasicCredentials = Depends(security)):
    username = credentials.username
    password = credentials.password
    
    if username == "admin" and password == "4dm1N" :
        return True
    else:
        raise HTTPException(status_code=401, detail="ONLY FOR ADMINS")
    




# subjects_list
subjects = set(row[1].lower() for row in sheet.iter_rows(min_row=2, values_only=True))

# types list
types_list = set(row[2].lower() for row in sheet.iter_rows(min_row=2, values_only=True))



@api.get('/', tags=["home"])
def greetings():
    """
    return a welcome message
    """
    return 'Welcome to the test database!'

@api.get('/health', name="Check the health of the application", tags=["home", "health"])
def health_check():
    """
    to check wether or not the api is running
    """
    return {'message': 'API is up and running'}


@api.get('/test', tags=["home"])
def feedback():

    """
    explain how to proceed to the exam equestions retrieval
    """
    return 'Please, choose the question type and subjects using /test/type:str/subject:str or /test/type:str/subject:List'

@api.get('/test/{type}/{subject}/{q_num}', name="get the questions based on input criteria")
def gen_exam(type: str, subject: str, q_num: int, verified: bool = Depends(verify_user)):
    """
    enter the type as str, subject(s) as str with only comma separed(no space!) and number of questions.\
    
    Example: validation test/utomation,datastreaming/10.\
    
    The only valid question numbers are 5,10,20, and then verify your username and password
    """

    #
    type = type.lower()
    subject = subject.lower()
    q = [5, 10, 20]
    if q_num not in q:
        raise HTTPException(status_code=404, detail="Choose 5, 10 or 20 questions Only!")

    if type not in types_list:
        raise HTTPException(status_code=400, detail="Type is not correct")

    if subject == "all" or subject is None:
        questions = [row for row in sheet.iter_rows(min_row=2, values_only=True)
                     if type == row[2].lower()]
        random.shuffle(questions)
    else:
        subjects_list = [subj.lower() for subj in subject.split(",")]
        invalid_subjects = [subj for subj in subjects_list if subj not in subjects]

        if invalid_subjects:
            raise HTTPException(status_code=400,\
                                 detail=f"Invalid subjects: {invalid_subjects}")

        questions = [row for row in sheet.iter_rows(min_row=2, values_only=True)
                     if type == row[2].lower() and row[1].lower() in subjects_list]
        
        random.shuffle(questions)


# Exam creation based on the number of questions input
    exam = []
    for quest in questions[:q_num]:
        test = {
            'Question': quest[0],
            'Subject': quest[1],
            'Choice A': quest[4],
            'Choice B': quest[5],
            'Choice C': quest[6]
        }
        if quest[7] is not None:
            test['choice D'] = quest[7]
        exam.append(test)

    if not exam:
        raise HTTPException(status_code=404, detail="No questions found for the specified criteria")

    return exam

# class represnts the new qustion to be added
class Quiz(BaseModel):
    question: str 
    subject: str
    use: str
    correct: str
    responseA: str
    responseB: str
    responseC: str
    responseD: Optional[str]


@api.put('/update', name="Add new question by an admin")
def put_quiz(q: Quiz, verified: bool = Depends(verify_admin)):
    """
    Add new question based after verified admin username and password
    """

    new_question = [
        q.question,
        q.subject,
        q.use,
        q.correct,
        q.responseA,
        q.responseB,
        q.responseC,
        q.responseD or ""  # Set to an empty string if responseD is None
    ]
    
    sheet.append(new_question)
    db.save('questions_en.xlsx')  # Save the updated Excel file
    
    return {"message": "Question added successfully"}





